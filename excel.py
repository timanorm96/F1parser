"""
Excel exporter using openpyxl.
Produces a multi-sheet workbook with styled tables and embedded charts.
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers as num_fmt
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ── Colour palette (F1-ish) ───────────────────────────────────────────────────
C_RED_DARK   = "C8102E"   # F1 red
C_RED_LIGHT  = "F5C6CB"
C_GOLD       = "FFD700"
C_SILVER     = "C0C0C0"
C_BRONZE     = "CD7F32"
C_HEADER_BG  = "1A1A2E"  # near-black navy
C_HEADER_FG  = "FFFFFF"
C_STRIPE_ODD = "F8F9FA"
C_STRIPE_EVEN= "FFFFFF"
C_GREEN      = "D4EDDA"
C_BORDER     = "DEE2E6"

THIN = Side(style="thin", color=C_BORDER)
THICK = Side(style="medium", color="AAAAAA")


def _border(left=True, right=True, top=True, bottom=True):
    return Border(
        left=THIN if left else None,
        right=THIN if right else None,
        top=THIN if top else None,
        bottom=THIN if bottom else None,
    )


def _header_style(cell):
    cell.font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=10)
    cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, freeze: bool = True) -> int:
    """Write a DataFrame to ws starting at start_row. Returns next free row."""
    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name).replace("_", " ").title())
        _header_style(cell)

    # Data rows
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = _border()
            # Zebra striping
            fill_color = C_STRIPE_ODD if (row_idx - start_row) % 2 == 0 else C_STRIPE_EVEN
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center" if isinstance(value, (int, float)) else "left")

    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto column width (capped at 40)
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if not df.empty else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    return start_row + len(df) + 2


def _highlight_podiums(ws, start_row: int, pos_col: int, n_rows: int):
    """Colour rows by finishing position (1=gold, 2=silver, 3=bronze)."""
    colours = {1: C_GOLD, 2: C_SILVER, 3: C_BRONZE}
    max_col = ws.max_column
    for r in range(start_row + 1, start_row + 1 + n_rows):
        pos_cell = ws.cell(row=r, column=pos_col)
        colour = colours.get(pos_cell.value)
        if colour:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=colour)


def _add_bar_chart(ws, title: str, data_col: int, label_col: int,
                   start_row: int, n_rows: int, anchor: str):
    chart = BarChart()
    chart.type = "bar"
    chart.title = title
    chart.y_axis.title = ""
    chart.x_axis.title = ""
    chart.style = 10
    chart.width = 20
    chart.height = 14

    data = Reference(ws, min_col=data_col, min_row=start_row, max_row=start_row + n_rows)
    cats = Reference(ws, min_col=label_col, min_row=start_row + 1, max_row=start_row + n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = C_RED_DARK
    ws.add_chart(chart, anchor)


def _add_line_chart(ws, title: str, data_cols: list[int], labels: list[str],
                    start_row: int, n_rows: int, anchor: str):
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = 24
    chart.height = 14

    colours_hex = [C_RED_DARK, "185FA5", "3B6D11", "854F0B", "534AB7"]
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + n_rows)

    for i, (col_idx, label) in enumerate(zip(data_cols, labels)):
        data = Reference(ws, min_col=col_idx, min_row=start_row, max_row=start_row + n_rows)
        chart.add_data(data, titles_from_data=True)
        series = chart.series[i]
        series.title = label
        series.graphicalProperties.line.solidFill = colours_hex[i % len(colours_hex)]
        series.graphicalProperties.line.width = 20000

    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def export_excel(
    output_path: str | Path,
    race_results: pd.DataFrame | None = None,
    qualifying: pd.DataFrame | None = None,
    driver_standings: pd.DataFrame | None = None,
    constructor_standings: pd.DataFrame | None = None,
    season_summary: pd.DataFrame | None = None,
    driver_comparison: pd.DataFrame | None = None,
    points_progression: pd.DataFrame | None = None,
    season: int = 0,
    race_name: str = "",
) -> Path:
    """
    Build a styled F1 analytics workbook.

    Pass whichever DataFrames you have; sheets are skipped if None.
    """
    output_path = Path(output_path)
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Sheet 1: Race Results ─────────────────────────────────────────────────
    if race_results is not None and not race_results.empty:
        ws = wb.create_sheet("Race Results")
        ws.sheet_view.showGridLines = False
        cols = [c for c in ["position", "driver_name", "team", "grid",
                             "laps", "points", "time_sec", "status",
                             "fastest_lap_time"] if c in race_results.columns]
        df = race_results[cols].copy()
        next_row = _write_df(ws, df, start_row=1)
        _highlight_podiums(ws, 1, list(df.columns).index("position") + 1, len(df))

        # Chart: points per driver
        pts_col = list(df.columns).index("points") + 1
        name_col = list(df.columns).index("driver_name") + 1
        _add_bar_chart(ws, f"Points — {race_name or 'Race'}", pts_col, name_col, 1, len(df), "K2")

    # ── Sheet 2: Qualifying ───────────────────────────────────────────────────
    if qualifying is not None and not qualifying.empty:
        ws = wb.create_sheet("Qualifying")
        ws.sheet_view.showGridLines = False
        cols = [c for c in ["position", "driver_name", "team",
                             "q1_sec", "q2_sec", "q3_sec", "best_time_sec"] if c in qualifying.columns]
        _write_df(ws, qualifying[cols], start_row=1)

    # ── Sheet 3: Driver Standings ──────────────────────────────────────────────
    if driver_standings is not None and not driver_standings.empty:
        ws = wb.create_sheet("Driver Standings")
        ws.sheet_view.showGridLines = False
        cols = [c for c in ["position", "driver_name", "team", "points",
                             "wins", "gap_to_leader"] if c in driver_standings.columns]
        df = driver_standings[cols].copy()
        next_row = _write_df(ws, df, start_row=1)
        _highlight_podiums(ws, 1, 1, min(3, len(df)))

        # Conditional data bar on points column
        pts_col_letter = get_column_letter(list(df.columns).index("points") + 1)
        pts_range = f"{pts_col_letter}2:{pts_col_letter}{len(df) + 1}"
        ws.conditional_formatting.add(pts_range,
            DataBarRule(start_type="min", end_type="max",
                        color=C_RED_DARK))

        _add_bar_chart(ws, "Championship Points", list(df.columns).index("points") + 1,
                       list(df.columns).index("driver_name") + 1, 1, len(df), "I2")

    # ── Sheet 4: Constructor Standings ────────────────────────────────────────
    if constructor_standings is not None and not constructor_standings.empty:
        ws = wb.create_sheet("Constructor Standings")
        ws.sheet_view.showGridLines = False
        cols = [c for c in ["position", "team", "points", "wins",
                             "gap_to_leader"] if c in constructor_standings.columns]
        df = constructor_standings[cols].copy()
        _write_df(ws, df, start_row=1)
        _highlight_podiums(ws, 1, 1, min(3, len(df)))

    # ── Sheet 5: Season Summary ───────────────────────────────────────────────
    if season_summary is not None and not season_summary.empty:
        ws = wb.create_sheet("Season Summary")
        ws.sheet_view.showGridLines = False
        _write_df(ws, season_summary, start_row=1)

    # ── Sheet 6: Driver Comparison ────────────────────────────────────────────
    if driver_comparison is not None and not driver_comparison.empty:
        ws = wb.create_sheet("Driver Comparison")
        ws.sheet_view.showGridLines = False
        _write_df(ws, driver_comparison, start_row=1, freeze=False)

        # Highlight "Better" column
        better_col_idx = list(driver_comparison.columns).index("Better") + 1 if "Better" in driver_comparison.columns else None
        if better_col_idx:
            for r in range(2, len(driver_comparison) + 2):
                cell = ws.cell(row=r, column=better_col_idx)
                if cell.value and cell.value != "—":
                    cell.fill = PatternFill("solid", fgColor=C_GREEN)

    # ── Sheet 7: Points Progression ──────────────────────────────────────────
    if points_progression is not None and not points_progression.empty:
        ws = wb.create_sheet("Points Progression")
        ws.sheet_view.showGridLines = False

        # Pivot: rounds as rows, drivers as columns
        if "cumulative_points" in points_progression.columns:
            pivot = points_progression.pivot_table(
                index=["round", "race_name"],
                columns="driver_name",
                values="cumulative_points",
                aggfunc="first"
            ).reset_index()
            pivot.columns.name = None

            _write_df(ws, pivot, start_row=1, freeze=False)

            drivers = [c for c in pivot.columns if c not in ("round", "race_name")]
            data_cols = [list(pivot.columns).index(d) + 1 for d in drivers]
            _add_line_chart(ws, "Points Progression", data_cols, drivers, 1, len(pivot), "A" + str(len(pivot) + 5))

    wb.save(output_path)
    return output_path
